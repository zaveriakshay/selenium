import openpyxl, pprint


class Beneficiary:

    def __index__(self):
        self.beneCode = ""
        self.beneAmount = ""


class Payment:

    def __init__(self):
        self.merchantCode = ""
        self.merchantRef = ""
        self.paymentAmount = ""
        self.paymentRemark = ""
        self.paymentMode = ""
        self.walletId = ""
        self.password = ""
        self.beneficiaryList = {}
        self.ccdHolderName = ""
        self.ccdExpiryMonth = ""
        self.ccdExpiryYear = ""
        self.ccdNumber = ""
        self.ccdCvv = ""
        self.ccdEmail = ""
        self.ddbBank = ""
        self.ddbName = ""
        self.ddbEmail = ""

        self.noqodiRef = ""
        self.paymentResponseXML = ""
        self.paymentErrorCode = ""

        self.sequence = ""
        self.isRefund = ""
        self.isEOD = ""
        self.isBene = ""
        self.Type = 0


class Refund:

    def __init__(self):
        self.merchantCode = ""
        self.sequence = ""
        self.paymentSequence = ""
        self.noqodiRef = ""
        self.merchantRef = ""
        self.refundAmount = ""
        self.refundRemark = ""
        self.beneficiaryList = {}
        self.refundSourceType = ""
        self.refundWalletId = ""
        self.refundIBANBankCountry = ""
        self.refundIBANBeneName = ""
        self.refundIBANNumber = ""

        self.noqodiRevRef = ""
        self.reversalResponseXML = ""
        self.reversalErrorCode = ""

        self.isEOD = ""
        self.isBene = ""
        self.Type = 0
        self.Filler3 = ""


class FIPaymentMessage:

    def __init__(self):
        self.sequence = ""
        self.FICode = ""
        self.FITransactionID = ""
        self.PaymentMode = ""
        self.PayRequestNo = ""
        self.Amount = ""
        self.Currency = ""
        self.ResponseXML = ""
        self.ResponseCode = ""
        self.noqodiRef = ""
        self.isEOD = ""
        self.PaymentDate = ""
        self.TransactionXML = ""
        self.Type = 0


class FIEOD:

    def __init__(self):
        self.sequence = ""
        self.FICode = ""
        self.FITransactionID = ""
        self.TotalTransactions = ""
        self.TotalAmount = ""
        self.ResponseXML = ""
        self.ResponseCode = ""
        self.TransactionXML = ""
        self.Type = 0
        self.noqodiRef = ""


def excelToExecutable(excelPath, sheetName):
    payment = Payment()
    wb = openpyxl.load_workbook(excelPath)
    print(wb.sheetnames)
    sheet_test = wb[sheetName]
    executableList = []
    executable = {}
    templatePayment = {}
    templateRefund = {}
    templateFIPaymentMessage = {}
    templateFIEOD = {}

    for rowIndex in range(sheet_test.min_row + 1, sheet_test.max_row + 1):
        if sheet_test.cell(row=rowIndex, column=1).value == "PAYMENT":
            executable = Payment()
            templatePayment = Payment()
        elif sheet_test.cell(row=rowIndex, column=1).value != None and "REFUND" in sheet_test.cell(row=rowIndex,
                                                                                                   column=1).value:
            executable = Refund()
            templateRefund = Refund()
        elif sheet_test.cell(row=rowIndex, column=1).value == "FIPAYMENTMESSAGE":
            executable = FIPaymentMessage()
            templateFIPaymentMessage = FIPaymentMessage()
        elif sheet_test.cell(row=rowIndex, column=1).value == "FIEOD":
            executable = FIEOD()
            templateFIEOD = FIEOD()

        executableList.append(executable)
        for columnIndex in range(sheet_test.min_column, sheet_test.max_column + 1):
            attr_column_mapping = str(sheet_test.cell(row=1, column=columnIndex).value)
            attr_column_value = str(sheet_test.cell(row=rowIndex, column=columnIndex).value)
            if isinstance(executable, Payment):
                setattr(templatePayment, attr_column_mapping, columnIndex)
            elif isinstance(executable, Refund):
                setattr(templateRefund, attr_column_mapping, columnIndex)
            elif isinstance(executable, FIPaymentMessage):
                setattr(templateFIPaymentMessage, attr_column_mapping, columnIndex)
            elif isinstance(executable, FIEOD):
                setattr(templateFIEOD, attr_column_mapping, columnIndex)

            print(attr_column_mapping + ":" + attr_column_value)
            if "bene" in attr_column_mapping:
                print(attr_column_mapping)
                executable.beneficiaryList[attr_column_mapping] = attr_column_value
            setattr(executable, attr_column_mapping, attr_column_value)

    return {"templateRefund": templateRefund, "templatePayment": templatePayment, "templateFIEOD": templateFIEOD,
            "templateFIPaymentMessage": templateFIPaymentMessage, "executableList": executableList}


def saveObjectToExcel(excelPath, sheetName, executable_list_, template_payment_, template_refund_,
                      templateFIPaymentMessage, templateFIEOD):
    wb = openpyxl.load_workbook(excelPath)
    # wb.create_sheet(index=0, title='PaymentsOutput')
    sheet_test = wb[sheetName]
    rowIndex = 0
    for executable in executable_list_:
        for propertyName in dir(executable):
            try:
                if not propertyName.startswith('__'):
                    if isinstance(executable, Payment):
                        property_index_ = getattr(template_payment_, propertyName)
                        rowIndex = getattr(executable, "sequence")
                    elif isinstance(executable, Refund):
                        property_index_ = getattr(template_refund_, propertyName)
                        print(":propertyName:>>" + str(propertyName) + "property_index_:>>" + str(property_index_))
                        rowIndex = getattr(executable, "sequence")
                        print(":propertyName:>>" + str(propertyName) + "rowIndex:>>" + str(rowIndex))
                    elif isinstance(executable, FIPaymentMessage):
                        property_index_ = getattr(templateFIPaymentMessage, propertyName)
                        print(":propertyName:>>" + str(propertyName) + "property_index_:>>" + str(property_index_))
                        rowIndex = getattr(executable, "sequence")
                        print(":propertyName:>>" + str(propertyName) + "rowIndex:>>" + str(rowIndex))
                    elif isinstance(executable, FIEOD):
                        property_index_ = getattr(templateFIEOD, propertyName)
                        print(":propertyName:>>" + str(propertyName) + "property_index_:>>" + str(property_index_))
                        rowIndex = getattr(executable, "sequence")
                        print(":propertyName:>>" + str(propertyName) + "rowIndex:>>" + str(rowIndex))

                    print(str(rowIndex) + "<<>>" + propertyName + "<<>>" + str(property_index_) + "<<>>" + str(
                        getattr(executable, propertyName)))
                    if "beneficiaryList" not in propertyName:
                        sheet_test.cell(row=int(rowIndex) + 1, column=int(str(property_index_))).value = str(
                            getattr(executable, propertyName))
            except Exception as exception:
                print(exception)
                pass
    wb.save(excelPath)


'''paymentOrRefund = excelToExecutable()
print(paymentOrRefund)
template_refund_ = paymentOrRefund["templateRefund"]
template_payment_ = paymentOrRefund["templatePayment"]
executable_list_ = paymentOrRefund["executableList"]
saveObjectToExccel(executable_list_, template_payment_, template_refund_)'''
